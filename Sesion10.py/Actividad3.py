import openpyxl 

libro = openpyxl.load_workbook(r"C:\Users\meryc\Desktop\mi_libro.xlsx")

hoja = libro["Ventas"]

datos_leidos = []

for fila in hoja.iter_rows(min_row=1, max_row=hoja.max_row, min_col=1, max_col=hoja.max_column, values_only=True): 

    datos_leidos.append(list(fila))

for fila in datos_leidos: 
    print(fila)