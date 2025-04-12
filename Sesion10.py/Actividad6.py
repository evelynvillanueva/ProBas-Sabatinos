import openpyxl 

libro = openpyxl.load_workbook(r"C:\Users\meryc\Desktop\mi_libro.xlsx")

if "Inventario" not in libro.sheetnames: 
    hoja_inventario = libro.create_sheet("Inventario")
else: 
    hoja_inventario = libro["Inventario"]

inventario = [
    ["Producto", "Stock"], 
    ["Manzanas", 100],
    ["Naranjas", 80]
]

for i, fila in enumerate(inventario, start=1):
    for j, valor in enumerate(fila, start=1):
        hoja_inventario.cell(row=i, column=j, value=valor)

libro.save(r"C:\Users\meryc\Desktop\mi_libro.xlsx")

hoja_ventas = libro["Ventas"]
ventas_dict = {}

for i in range(2, hoja_ventas.max_row + 1):
    producto = hoja_ventas.cell(row=i, column=1).value
    cantidad = hoja_ventas.cell(row=i, column=2).value
    precio = hoja_ventas.cell(row=i, column=3).value
    total = hoja_ventas.cell(row=i, column=4).value
    ventas_dict[producto] = {"Cantidad": cantidad, "Precio": precio, "Total": total}

hoja_inventario = libro["Inventario"]
inventario_dict = {}

for i in range(2, hoja_inventario.max_row + 1):
    producto = hoja_inventario.cell(row=i, column=1).value
    stock = hoja_inventario.cell(row=i, column=2).value
    inventario_dict[producto] = {"Stock": stock}

print("Diccionario de Ventas:")
print(ventas_dict)

print("\nDiccionario de Inventario:")
print(inventario_dict)