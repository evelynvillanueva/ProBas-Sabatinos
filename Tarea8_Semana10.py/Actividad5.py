import openpyxl 

libro = openpyxl.load_workbook(r"C:\Users\meryc\Desktop\mi_libro.xlsx")

hoja = libro["Ventas"]

hoja.cell(row=1, column=4, value="Total")

for i in range(2, hoja.max_row + 1): 

    hoja.cell(row=i, column=4, value=f"=B{i}*C{i}")

libro.save(r"C:\Users\meryc\Desktop\mi_libro.xlsx")

print("La columna 'Total' se ha agregado exitosamente.")