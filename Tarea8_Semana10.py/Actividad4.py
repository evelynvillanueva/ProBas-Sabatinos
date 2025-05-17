import openpyxl 

libro = openpyxl.load_workbook(r"C:\Users\meryc\Desktop\mi_libro.xlsx")

hoja = libro["Ventas"]

for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=1, max_col=hoja.max_column): 

    if fila[0].value == "Manzanas":
        fila[2].value = 0.55
        break 

libro.save(r"C:\Users\meryc\Desktop\mi_libro.xlsx")

print("El precio de las 'Manzanas' se ha modificado exitosamente.")